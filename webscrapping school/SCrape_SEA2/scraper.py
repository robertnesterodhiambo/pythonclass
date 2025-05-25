import requests
from datetime import datetime
import logging
import time
import traceback
from openpyxl import load_workbook

logging.basicConfig(
    filename=f"logs_{int(time.time())}.log",
    format="%(asctime)s - %(levelname)s [%(filename)s:%(lineno)s - %(funcName)s() ] %(message)s",
    filemode="w",
)
logger = logging.getLogger()
logger.setLevel(logging.INFO)

file_path = "/home/dragon/DATA/SEACO.xlsx"  # Source of units numbers
output_filename = "seaco.csv"  # Output file


def read_excel_as_chunks(file_path, chunk_size=25):
    "Read the units number from the excel file and return a list of list 25 numbers each"
    # Load the workbook and select the active worksheet
    workbook = load_workbook(file_path)
    sheet = workbook.active

    # Gather all values from the sheet into a flat list
    all_values = []
    for col in sheet.iter_cols(min_row=2, values_only=True):  # Skip header row
        all_values.extend(filter(None, col))  # Add non-None values
    print(f"Number of units: {len(all_values)}")
    print(f"Number of unique units: {len(set(all_values))}")
    # return all_values
    # Split the flat list into chunks of size chunk_size
    chunks = [
        all_values[i : i + chunk_size] for i in range(0, len(all_values), chunk_size)
    ]
    return chunks


headers = {
    "accept": "application/atomsvc+xml;q=0.8, application/json;odata=fullmetadata;q=0.7, application/json;q=0.5, */*;q=0.1",
    "accept-language": "fr-FR,fr;q=0.9,en-US;q=0.8,en;q=0.7,ar;q=0.6",
    "content-type": "application/json; charset=utf-8",
    "cookie": "sap-usercontext=sap-client=100; oucqsvzubqrobvzdoredfoaeacoassxuzatwuyx=GET#MIICIAYJKoZIhvcNAQc...",
    "dataserviceversion": "2.0",
    "maxdataserviceversion": "3.0",
    "priority": "u=1, i",
    "referer": "https://seaweb.seacoglobal.com/sap/bc/ui5_ui5/sap/zseaco_ue17/index.html",
    "sap-passport": "2A54482A0300E60000756E64657465726D696E6564202020202020202020202020202020202020202000005341505F453245...",
    "sec-ch-ua": '"Google Chrome";v="131", "Chromium";v="131", "Not_A Brand";v="24"',
    "sec-ch-ua-mobile": "?0",
    "sec-ch-ua-platform": '"Windows"',
    "sec-fetch-dest": "empty",
    "sec-fetch-mode": "cors",
    "sec-fetch-site": "same-origin",
    "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36",
    "x-csrf-token": "Fetch",
    "x-requested-with": "XMLHttpRequest",
}


def split_equnr(lst):
    # Ensure the list contains 25 elements or fewer
    if len(lst) > 25:
        raise ValueError("The input list must contain 25 elements or fewer.")

    # Split the list into two parts
    midpoint = (len(lst) + 1) // 2  # Ensures more elements go to Equnr if odd length
    equnr = lst[:midpoint]  # First part
    equnr1 = lst[midpoint:]  # Second part

    return equnr, equnr1


def writeCsv(row, output_filename):
    headers = [
        "Unit Number",
        "Unit Type",
        "Status",
        "Customer",
        "On/Off hire Date",
        "On/Off hire city",
        "Year of Manuf",
    ]
    headers = ";".join(headers)
    row = ";".join(row)
    try:
        with open(output_filename, "a", encoding="utf-8") as f:
            f.write(row + "\n")
    except Exception as e:
        logger.log(traceback.format_exc())


def makeRequest(url, params=None, headers=headers):
    "Make a get request and retry if response is anything else but 200"
    try:
        response = requests.get(url, params=params, headers=headers)
        if response.status_code == 200:
            results = response.json()["d"]["results"]
            return results
        elif response.status_code == 429:  # too many requests
            logger.error(f"{response.status_code}, {url}")
            ban_time = int(response.headers["Retry-After"])
            logger.error(f"Sleeping for {ban_time} seconds...")
            time.sleep(ban_time)
            return makeRequest(url, params=params, headers=headers)
        elif response.status_code == 502:  # bad gateway
            logger.error(f"{response.status_code}, {url}")
            time.sleep(1)
            return makeRequest(url, params=params, headers=headers)
        else:
            logger.info(f"{response.status_code}, {url}")
            return makeRequest(url, params=params, headers=headers)
    except requests.exceptions.JSONDecodeError:
        logger.error(traceback.format_exc())
        logger.error(response.text)
        logger.error(params)
    except Exception as e:  # network error
        logger.error(traceback.format_exc())
        time.sleep(10)  # sleep 10 seconds before retry
        return makeRequest(url, params=params, headers=headers)
    return response


def getUnitStatus(list_numbers=[]):

    url = "https://seaweb.seacoglobal.com/sap/opu/odata/sap/ZNW_SEACO_PUBLIC_UE_PGW_1_SRV/Unit_Status_Multiple"
    equnr, equnr1 = split_equnr(list_numbers)
    str_numbers = ",".join(equnr).upper()
    str_numbers1 = ",".join(equnr1).upper()
    params = {
        "$filter": f"Equnr eq '{str_numbers},' and Equnr1 eq '{str_numbers1},'",
        "saml2": "disabled",
    }
    results = makeRequest(url, params)
    list_units = []
    for result in results:
        Message = result["Message"]
        if Message == "Unit not found":
            logger.info(f"Not found: {result}")
        UnitNumber = result["UnitNumber"]
        ProductType = result["ProductType"]
        Status = result["Status"]
        Customer = result["Customer"]
        HireDate = result["HireDate"]
        City = result["City"]
        if HireDate != "/Date(253392451200000)/":  # this ts mean no date
            timestamp = HireDate.replace("/Date(", "").replace(")/", "")
            try:
                dt_object = datetime.fromtimestamp(float(timestamp) / 1000).date()
            except:
                logger.error(traceback.format_exc())
                logger.error(UnitNumber)
        else:
            dt_object = ""
        row = [UnitNumber, ProductType, Status, Customer, str(dt_object), City]
        list_units.append(row)
    return list_units


def getEquipementStatus(list_numbers):
    url = "https://seaweb.seacoglobal.com/sap/opu/odata/sap/ZNW_SEACO_PUBLIC_UE_PGW_1_SRV/Equipment_Status"
    equnr, equnr1 = split_equnr(list_numbers)
    str_numbers = ",".join(equnr).upper()
    str_numbers2 = ",".join(equnr1).upper()
    params = {
        "$filter": f"IUnitNo eq '{str_numbers},' and IUnitNo1 eq '{str_numbers2},'",
        "saml2": "disabled",
    }
    results = makeRequest(url, params=params, headers=headers)
    list_equipements = []
    for result in results:
        Message = result["Message"]
        if Message == "Unit not found":
            logger.info(f"Not found: {result}")
        Manmonth = result["Manmonth"]
        ProductType = result["Manyear"]
        year_man = f"{Manmonth}/{ProductType}"
        list_equipements.append(year_man)
    return list_equipements


def removeDuplicate(file_name, output_file):
    "Remove duplicate and the '/' in manf. year if empty"
    with open(file_name, "r", encoding="utf-8") as f:
        lines = f.readlines()

    with open(output_file, "w", encoding="utf-8") as f:
        for line in lines:
            line = line.replace(";/", ";")
            f.write(line)


def getAllUnits():
    chunks = read_excel_as_chunks(file_path)

    for chunk in chunks:
        print(f"Getting chunk {chunks.index(chunk)} from {len(chunks)}")
        list_units = getUnitStatus(chunk)
        list_equipements = getEquipementStatus(chunk)
        for i in range(len(list_units)):
            list_units[i].append(list_equipements[i])
            writeCsv(list_units[i], output_filename)


if __name__ == "__main__":
    getAllUnits()
    # removeDuplicate("seaco.csv", "full_seaco.csv")
